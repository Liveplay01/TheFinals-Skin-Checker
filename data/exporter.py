import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


RARITY_HEX = {
    "COMMON":    "9E9E9E",
    "RARE":      "2196F3",
    "EPIC":      "9C27B0",
    "LEGENDARY": "FFC107",
    "MYTHIC":    "E91E63",
}

HEADER_FILL = PatternFill(start_color="161B22", end_color="161B22", fill_type="solid")
HEADER_FONT = Font(color="E8C547", bold=True, name="Segoe UI", size=11)


def export_xlsx(skins: list, path: str) -> str:
    """
    Export detected skins to an Excel file.
    skins: list of DetectedSkin objects.
    Returns the path written.
    """
    rows = []
    for s in skins:
        rows.append({
            "Skin Name":     s.name,
            "Weapon":        s.weapon,
            "Build":         s.build,
            "Rarity":        s.rarity,
            "Source":        s.source,
            "Obtainable":    "Yes" if s.obtainable else "No",
            "Season":        s.season or "-",
            "Cost (Multibucks)": s.cost if s.cost > 0 else "-",
            "First Detected": s.first_seen,
        })

    df = pd.DataFrame(rows)
    df.sort_values(
        by=["Rarity", "Weapon", "Skin Name"],
        key=lambda col: col.map(lambda v: (["MYTHIC", "LEGENDARY", "EPIC", "RARE", "COMMON"].index(v) if v in ["MYTHIC", "LEGENDARY", "EPIC", "RARE", "COMMON"] else 99))
        if col.name == "Rarity" else col,
        inplace=True,
    )

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Detected Skins")
        ws = writer.sheets["Detected Skins"]

        # Style header row
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Color rarity column cells
        rarity_col_idx = df.columns.get_loc("Rarity") + 1
        for row_idx in range(2, len(df) + 2):
            cell = ws.cell(row=row_idx, column=rarity_col_idx)
            rarity = cell.value
            hex_color = RARITY_HEX.get(rarity, "FFFFFF")
            cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
            brightness = sum(int(hex_color[i:i+2], 16) for i in (0, 2, 4)) / 3
            cell.font = Font(
                color="000000" if brightness > 128 else "FFFFFF",
                bold=True,
                name="Segoe UI",
                size=10,
            )
            cell.alignment = Alignment(horizontal="center")

        # Thin border for all cells
        thin = Side(style="thin", color="2D333B")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for row in ws.iter_rows(min_row=1, max_row=len(df) + 1, max_col=len(df.columns)):
            for cell in row:
                cell.border = border

        # Auto-size columns
        for col_idx, col_name in enumerate(df.columns, start=1):
            max_len = max(
                len(str(col_name)),
                *(len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(2, len(df) + 2))
            ) + 4
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len, 40)

        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "A2"

    return path
